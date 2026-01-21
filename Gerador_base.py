import os
import re
import threading
import tkinter as tk
from tkinter import ttk, messagebox, filedialog

import pandas as pd
import mysql.connector

from openpyxl.utils import get_column_letter


# =========================
# CONFIG / CONSTANTES
# =========================
CRED_FILE_PATH = r"\\fs01\ITAPEVA ATIVAS\DADOS\SA_Credencials_Copia.txt"

CARTEIRAS = [
    ("517 Itapeva Autos", 517),
    ("518 DivZero", 518),
    ("519 Cedidas", 519),
]

TEL_LIMIT_FIXO = 7

# (Opção 1) Histórico: qual coluna em hist_tb referencia cadastros_tb.cod_cad?
# No seu SQL original está como "his.cod_cli". Se no seu banco for diferente, ajuste aqui.
HIST_CAD_REF_COL = "cod_cli"


# =========================
# Credenciais (arquivo)
# =========================
def parse_credentials_from_file(path: str) -> dict:
    if not os.path.exists(path):
        raise FileNotFoundError(f"Arquivo de credenciais nao encontrado: {path}")

    with open(path, "r", encoding="utf-8", errors="ignore") as f:
        content = f.read()

    def pick(key: str):
        m = re.search(rf"^{key}\s*=\s*(.+?)\s*$", content, re.MULTILINE)
        if not m:
            return None
        raw = m.group(1).split("#", 1)[0].strip()
        return raw.strip().strip('"').strip("'")

    host = pick("GECOBI_HOST")
    user = pick("GECOBI_USER")
    passwd = pick("GECOBI_PASS")
    db = pick("GECOBI_DB")
    port = pick("GECOBI_PORT")

    missing = [k for k, v in {
        "GECOBI_HOST": host,
        "GECOBI_USER": user,
        "GECOBI_PASS": passwd,
        "GECOBI_DB": db,
        "GECOBI_PORT": port
    }.items() if not v]

    if missing:
        raise ValueError("Chaves nao encontradas no arquivo de credenciais: " + ", ".join(missing))

    try:
        port_int = int(port)
    except ValueError:
        raise ValueError(f"GECOBI_PORT invalido no arquivo: {port!r}")

    return {
        "host": host,
        "user": user,
        "password": passwd,
        "database": db,
        "port": port_int
    }


# =========================
# Helpers
# =========================
def _is_valid_ymd(s: str) -> bool:
    if not s:
        return False
    return re.fullmatch(r"\d{4}-\d{2}-\d{2}", s) is not None


def _parse_money_br_or_plain(s: str) -> float:
    """
    Aceita:
      - "10000"
      - "10000.50"
      - "10.000"
      - "10.000,50"
      - "R$ 10.000,50"
    Retorna float.
    """
    if s is None:
        raise ValueError("Valor mínimo não informado.")
    raw = s.strip()
    if not raw:
        raise ValueError("Valor mínimo não informado.")

    raw = raw.replace("R$", "").replace(" ", "")
    # Se tem vírgula, assume formato BR (1.234,56)
    if "," in raw:
        raw = raw.replace(".", "").replace(",", ".")
    try:
        return float(raw)
    except ValueError:
        raise ValueError("Valor mínimo inválido. Ex.: 10000 ou 10.000,00")


def _write_excel_pretty(df: pd.DataFrame, path: str, sheet_name: str):
    """
    (Opção 10) Excel mais "profissional":
      - Freeze header
      - AutoFilter
      - Ajuste de largura (amostrando até 500 linhas)
      - Formatação numérica em colunas de valor
    """
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name=sheet_name)
        ws = writer.sheets[sheet_name]

        # Congela cabeçalho
        ws.freeze_panes = "A2"

        # Filtro automático
        ws.auto_filter.ref = ws.dimensions

        # Ajuste de largura (amostra para não ficar pesado)
        sample_rows = min(len(df), 500)
        for col_idx, col_name in enumerate(df.columns, start=1):
            max_len = len(str(col_name)) if col_name is not None else 10

            if sample_rows > 0:
                series = df[col_name].iloc[:sample_rows]
                for v in series:
                    if v is None:
                        continue
                    s = str(v)
                    if len(s) > max_len:
                        max_len = len(s)

            width = max(10, min(max_len + 2, 60))
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        # Formatação de colunas de valor (leve e útil)
        money_keywords = ("valor", "vlr", "divida", "acordo", "parc", "commission", "collected", "goal")
        for col_idx, col_name in enumerate(df.columns, start=1):
            col_lower = str(col_name).lower()

            # Se a coluna parece monetária E é numérica, aplica formato #,##0.00
            if any(k in col_lower for k in money_keywords) and pd.api.types.is_numeric_dtype(df[col_name]):
                for row in range(2, ws.max_row + 1):
                    ws.cell(row=row, column=col_idx).number_format = "#,##0.00"


# =========================
# Runner SQL (IN multi-carteiras)
# =========================
def build_sql_and_params(sql_template: str, carteiras: list[int], extra: dict | None = None) -> tuple[str, list]:
    """
    (Opção 2) Montagem determinística dos parâmetros:
      - Para cada ocorrência de {cod_cli} no template, adiciona a lista completa de carteiras
      - Depois adiciona os parâmetros de cauda (_tail_params)
    """
    if not carteiras:
        raise ValueError("Nenhuma carteira selecionada.")

    extra = extra or {}

    # Quantas vezes o template usa {cod_cli}
    cod_cli_occurrences = sql_template.count("{cod_cli}")
    if cod_cli_occurrences <= 0:
        raise ValueError("SQL template não contém {cod_cli}.")

    in_placeholders = ", ".join(["%s"] * len(carteiras))

    sql = sql_template.format(
        cod_cli=in_placeholders,
        tel_limit=TEL_LIMIT_FIXO,
        dt_ini_filter=extra.get("dt_ini_filter", ""),
        dt_fim_filter=extra.get("dt_fim_filter", ""),
        infoad_filter=extra.get("infoad_filter", ""),
        vlrparc_having=extra.get("vlrparc_having", ""),
        having_filter=extra.get("having_filter", ""),
        hist_cad_ref_col=extra.get("hist_cad_ref_col", HIST_CAD_REF_COL),  # (Opção 1)
    )

    tail_params = extra.get("_tail_params", None)
    if tail_params is None:
        tail_params = extra.get("_date_params", []) or []
    else:
        tail_params = tail_params or []

    params: list = []
    for _ in range(cod_cli_occurrences):
        params.extend(carteiras)
    params.extend(tail_params)

    return sql, params


def run_query(sql_template: str, carteiras: list[int], extra: dict | None = None) -> pd.DataFrame:
    creds = parse_credentials_from_file(CRED_FILE_PATH)
    sql, params = build_sql_and_params(sql_template, carteiras, extra=extra)

    conn = mysql.connector.connect(**creds)
    try:
        return pd.read_sql(sql, conn, params=params)
    finally:
        conn.close()


# =========================
# SQLs (SEUS - mantidos)
# =========================
SQL_EMAIL = """
SELECT nomecli, cpfcnpj, email
FROM cadastros_tb
WHERE cod_cli IN ({cod_cli})
  AND stcli <> 'INA';
"""

SQL_NOME_CPF = """
SELECT nomecli, cpfcnpj
FROM cadastros_tb
WHERE cod_cli IN ({cod_cli})
  AND stcli <> 'INA';
"""

SQL_TELEFONES_MELHOR_CONTATO = r"""
WITH telefones AS (
    SELECT
        cad.cod_cad,
        cad.nomecli AS nome,
        cad.cpfcnpj AS cpf,
        cad.nmcont,
        cad.cod_cli,
        cad.infoad AS portfolio,
        CONCAT(tel.dddfone, tel.telefone) AS telefone,
        tel.status,
        tel.obs,
        ROW_NUMBER() OVER (
            PARTITION BY cad.cod_cad
            ORDER BY FIELD(tel.status, 2, 4, 5, 6, 1, 0), tel.status
        ) AS num
    FROM cadastros_tb cad
    JOIN fones_tb tel
        ON tel.cod_cad = cad.cod_cad
    WHERE cad.cod_cli IN ({cod_cli})
      AND cad.stcli <> 'INA'
      AND (tel.status IN (2,4,5,6,1)
           OR (tel.obs NOT LIKE '%Descon%' AND tel.obs NOT LIKE '%incorret%'))
      AND CONCAT(tel.dddfone, tel.telefone) NOT REGEXP '([0-9])\\1{{5}}'
      AND LENGTH(CONCAT(tel.dddfone, tel.telefone)) >= 8
      AND CONCAT(tel.dddfone, tel.telefone) NOT LIKE '%X%'
),
filtrados AS (
    SELECT * FROM telefones WHERE num <= {tel_limit}
)
SELECT
    cod_cad, nome, cpf, nmcont, cod_cli, portfolio,

    MAX(CASE WHEN num = 1 THEN telefone END) AS Telefone1,
    MAX(CASE WHEN num = 1 THEN status   END) AS StatusTelefone1,
    MAX(CASE WHEN num = 1 THEN obs      END) AS ObsTelefone1,

    MAX(CASE WHEN num = 2 THEN telefone END) AS Telefone2,
    MAX(CASE WHEN num = 3 THEN telefone END) AS Telefone3,
    MAX(CASE WHEN num = 4 THEN telefone END) AS Telefone4,
    MAX(CASE WHEN num = 5 THEN telefone END) AS Telefone5,
    MAX(CASE WHEN num = 6 THEN telefone END) AS Telefone6,
    MAX(CASE WHEN num = 7 THEN telefone END) AS Telefone7
FROM filtrados
GROUP BY cod_cad, nome, cpf, nmcont, cod_cli, portfolio;
"""

SQL_ACORDOS_PA = r"""
WITH ranked AS (
    SELECT
        a.nmcont,
        a.cod_cli,
        a.cod_aco,
        a.data_cad,
        a.vlr_aco,
        a.qtd_p_aco,
        a.staco,
        ROW_NUMBER() OVER (PARTITION BY a.nmcont, a.cod_cli ORDER BY a.cod_aco DESC) AS rn
    FROM acordos_tb a
    WHERE a.cod_cli IN ({cod_cli})
      AND a.data_cad >= '2025-07-01'
)
SELECT
    nmcont,
    cod_cli,
    cod_aco,
    DATE_FORMAT(data_cad, '%d-%m-%Y') AS DataCriacaoGecobi,
    vlr_aco AS UltimoValorAcordado,
    qtd_p_aco,
    CASE WHEN qtd_p_aco = 1 THEN 'AVISTA' ELSE 'PARCELADO' END AS TipoAcordo,
    staco
FROM ranked
WHERE rn = 1
  AND staco IN ('P','A')
ORDER BY cod_aco DESC;
"""

# (Opção 1) CPC por período — JOIN parametrizável para ficar claro.
SQL_CPC_PERIODO = r"""
SELECT
    cad.cod_cad,
    cad.nomecli AS nome,
    cad.cpfcnpj AS cpf,
    cad.nmcont,
    MAX(his.data_at) AS dt_ultimo_cpc
FROM cadastros_tb cad
JOIN hist_tb his
    ON his.{hist_cad_ref_col} = cad.cod_cad
JOIN stcob_tb st
    ON st.st = his.ocorr
WHERE cad.cod_cli IN ({cod_cli})
  AND cad.stcli <> 'INA'
  AND st.bsc LIKE '%CPC%'
  {dt_ini_filter}
  {dt_fim_filter}
GROUP BY cad.cod_cad, cad.nomecli, cad.cpfcnpj, cad.nmcont;
"""

SQL_SEM_HIST_30D = r"""
SELECT
    cad.cod_cad,
    cad.nomecli AS nome,
    cad.cpfcnpj AS cpf,
    cad.nmcont,
    cad.cod_cli,
    cad.infoad AS portfolio
FROM cadastros_tb cad
WHERE cad.cod_cli IN ({cod_cli})
  AND cad.stcli <> 'INA'
  AND cad.cod_cad NOT IN (
      SELECT h.cod_cli
      FROM hist_tb h
      WHERE h.data_at >= NOW() - INTERVAL 30 DAY
        AND h.cod_usu <> '999'
      GROUP BY h.cod_cli
  );
"""

SQL_RECENTES = r"""
WITH
cpc_ultimo AS (
    SELECT
        st.bsc,
        cad.nmcont,
        MAX(his.data_at) AS dt_ultimo_cpc
    FROM cadastros_tb cad
    JOIN hist_tb his
        ON his.cod_cli = cad.cod_cad
    JOIN stcob_tb st
        ON st.st = his.ocorr
    WHERE cad.cod_cli IN ({cod_cli})
      AND cad.stcli <> 'INA'
    GROUP BY cad.nmcont
),
acordos_ranked AS (
    SELECT
        a.nmcont,
        a.cod_aco,
        a.data_aco,
        a.data_cad,
        a.vlr_aco,
        a.qtd_p_aco,
        a.staco,
        ROW_NUMBER() OVER (
            PARTITION BY a.nmcont
            ORDER BY a.cod_aco DESC
        ) AS rn_aco
    FROM acordos_tb a
    WHERE a.cod_cli IN ({cod_cli})
      AND a.data_cad >= '2025-07-01'
),
acordos_pagos AS (
    SELECT nmcont
    FROM acordos_ranked
    WHERE rn_aco = 1
      AND staco IN ('P','G','A')
),
acordos_ultimos AS (
    SELECT
        aco.nmcont,
        aco.vlr_aco AS UltimoValorAcordado,
        CASE
            WHEN aco.qtd_p_aco = 1 THEN 'AVISTA'
            WHEN aco.qtd_p_aco > 1 THEN 'PARCELADO'
            ELSE NULL
        END AS TipoAcordo,
        DATE_FORMAT(aco.data_cad, '%d-%m-%Y') AS DataCriacaoGecobi,
        CASE aco.staco
            WHEN 'A' THEN 'Em Acordo'
            WHEN 'E' THEN 'Exceção Rejeitada'
            WHEN 'G' THEN 'Pago'
            WHEN 'Q' THEN 'Quebrado'
            WHEN 'P' THEN 'Em Promessa'
            ELSE 'Sem Dados'
        END AS StatusUltimoAcordo,
        aco.staco
    FROM acordos_ranked aco
    WHERE aco.rn_aco = 1
      AND aco.staco IN ('P','G','A','Q','E')
),
valores AS (
    SELECT
        recc.nmcont,
        recc.cod_cli,
        GROUP_CONCAT(DISTINCT rec.fat_parc ORDER BY rec.fat_parc SEPARATOR ' || ') AS Contratos,
        GROUP_CONCAT(DISTINCT recc.char_5 ORDER BY recc.char_5 SEPARATOR ' || ') AS TipoProduto
    FROM rec_comp_tb recc
    LEFT JOIN receber_tb rec
        ON rec.nmcont = recc.nmcont
       AND rec.cod_cli = recc.cod_cli
    WHERE recc.cod_cli IN ({cod_cli})
      AND rec.fat_parc NOT LIKE '%ENTRADA%'
      AND recc.nmcont NOT IN (SELECT nmcont FROM acordos_pagos)
    GROUP BY recc.nmcont, recc.cod_cli
),
bens AS (
    SELECT
        ben.nmcont,
        ben.cod_cli,
        CONCAT(ben.marca, ' - ', ben.modelo) AS MarcaModelo,
        ben.placa,
        ben.cor,
        CONCAT(ben.anofab, '/', ben.anomodelo) AS AnoFabModelo,
        COUNT(DISTINCT ben.chassi) AS QtdGarantiasUnicas
    FROM bens_tb ben
    WHERE ben.cod_cli IN ({cod_cli})
    GROUP BY ben.nmcont, ben.cod_cli
),
telefones AS (
    SELECT
        cad.cod_cad AS cod_cad,
        cad.nomecli AS nome,
        cad.cpfcnpj AS cpf,
        cad.nmcont AS nmcont,
        cad.cod_cli AS cod_cli,
        MAX(recc.int_2) AS BindingID,
        DATE_FORMAT(nascto, '%d-%m-%Y') AS DataNascimento,
        cad.infoad AS Portfolio,
        CONCAT(dddfone,telefone) AS telefones,
        ROW_NUMBER() OVER (
            PARTITION BY tel.cod_cad
            ORDER BY FIELD(tel.status, 2, 4, 5, 6, 1, 0), tel.status
        ) AS num
    FROM cadastros_tb cad
    JOIN fones_tb tel
        ON tel.cod_cad = cad.cod_cad
    LEFT JOIN rec_comp_tb recc
        ON recc.nmcont = cad.nmcont
       AND cad.cod_cli = recc.cod_cli
    WHERE cad.cod_cli IN ({cod_cli})
      AND cad.data_cad = cad.data_arq
      AND cad.data_cad >= (curdate() - interval 2 month)
      AND CONCAT(dddfone, telefone) NOT REGEXP '([0-9])\\1{{5}}'
      AND cad.stcli <> 'INA'
    GROUP BY
        cad.cod_cad,
        cad.nomecli,
        cad.cpfcnpj,
        cad.nmcont,
        cad.cod_cli,
        nascto,
        cad.infoad,
        dddfone,
        telefone,
        tel.status
),
telefones_filtrados AS (
    SELECT * FROM telefones WHERE num <= {tel_limit}
),
telefones_final AS (
    SELECT
        cod_cad,
        nome,
        cpf,
        nmcont,
        cod_cli,
        BindingID,
        DataNascimento,
        Portfolio,
        MAX(CASE WHEN num = 1 THEN telefones END) AS Telefone1,
        MAX(CASE WHEN num = 2 THEN telefones END) AS Telefone2,
        MAX(CASE WHEN num = 3 THEN telefones END) AS Telefone3,
        MAX(CASE WHEN num = 4 THEN telefones END) AS Telefone4,
        MAX(CASE WHEN num = 5 THEN telefones END) AS Telefone5,
        MAX(CASE WHEN num = 6 THEN telefones END) AS Telefone6,
        MAX(CASE WHEN num = 7 THEN telefones END) AS Telefone7
    FROM telefones_filtrados
    GROUP BY
        cod_cad, nome, cpf, nmcont, cod_cli, BindingID, DataNascimento, Portfolio
)
SELECT
    t.cod_cad,
    t.nome,
    t.cpf,
    t.BindingID,
    t.DataNascimento,
    t.Portfolio,
    t.Telefone1, t.Telefone2, t.Telefone3, t.Telefone4, t.Telefone5, t.Telefone6, t.Telefone7,
    b.MarcaModelo,
    b.placa,
    b.cor,
    b.AnoFabModelo,
    b.QtdGarantiasUnicas,
    v.Contratos,
    v.TipoProduto,
    a.UltimoValorAcordado,
    a.TipoAcordo,
    a.DataCriacaoGecobi,
    a.StatusUltimoAcordo
FROM telefones_final t
LEFT JOIN bens b
    ON t.nmcont = b.nmcont
   AND t.cod_cli = b.cod_cli
JOIN valores v
    ON t.nmcont = v.nmcont
   AND t.cod_cli = v.cod_cli
LEFT JOIN acordos_ultimos a
    ON t.nmcont = a.nmcont
LEFT JOIN cpc_ultimo cpc
    ON t.nmcont = cpc.nmcont
WHERE t.cod_cli IN ({cod_cli});
"""

SQL_NUNCA = r"""
WITH
cpc_ultimo AS (
    SELECT
        cad.nmcont,
        MAX(his.data_at) AS dt_ultimo_cpc
    FROM cadastros_tb cad
    JOIN hist_tb his ON his.cod_cli = cad.cod_cad
    JOIN stcob_tb st ON st.st = his.ocorr
    WHERE cad.cod_cli IN ({cod_cli})
      AND cad.stcli <> 'INA'
      AND st.bsc LIKE '%%CPC%%'
    GROUP BY cad.nmcont
),
acordos_ranked AS (
    SELECT
        a.nmcont,
        a.cod_aco,
        a.data_aco,
        a.data_cad,
        a.vlr_aco,
        a.qtd_p_aco,
        a.staco,
        ROW_NUMBER() OVER (PARTITION BY a.nmcont ORDER BY a.cod_aco DESC) AS rn_aco
    FROM acordos_tb a
    WHERE a.cod_cli IN ({cod_cli})
      AND a.data_cad >= '2025-07-01'
),
acordos_pagos AS (
    SELECT nmcont
    FROM acordos_ranked
    WHERE rn_aco = 1
      AND staco IN ('P','G','A')
),
acordos_ultimos AS (
    SELECT
        aco.nmcont,
        aco.vlr_aco AS UltimoValorAcordado,
        CASE
            WHEN aco.qtd_p_aco = 1 THEN 'AVISTA'
            WHEN aco.qtd_p_aco > 1 THEN 'PARCELADO'
            ELSE NULL
        END AS TipoAcordo,
        DATE_FORMAT(aco.data_cad, '%d-%m-%Y') AS DataCriacaoGecobi,
        CASE aco.staco
            WHEN 'A' THEN 'Em Acordo'
            WHEN 'E' THEN 'Exceção Rejeitada'
            WHEN 'G' THEN 'Pago'
            WHEN 'Q' THEN 'Quebrado'
            WHEN 'P' THEN 'Em Promessa'
            ELSE 'Sem Dados'
        END AS StatusUltimoAcordo,
        aco.staco
    FROM acordos_ranked aco
    WHERE aco.rn_aco = 1
      AND aco.staco IN ('Q','E','P','G','A')
),
valores AS (
    SELECT
        recc.nmcont,
        recc.cod_cli,
        GROUP_CONCAT(DISTINCT rec.fat_parc ORDER BY rec.fat_parc SEPARATOR ' || ') AS Contratos,
        GROUP_CONCAT(DISTINCT recc.char_5 ORDER BY recc.char_5 SEPARATOR ' || ') AS TipoProduto
    FROM rec_comp_tb recc
    LEFT JOIN receber_tb rec
        ON rec.nmcont = recc.nmcont
       AND rec.cod_cli = recc.cod_cli
    WHERE recc.cod_cli IN ({cod_cli})
      AND rec.fat_parc NOT LIKE '%%ENTRADA%%'
      AND recc.nmcont NOT IN (SELECT nmcont FROM acordos_pagos)
    GROUP BY recc.nmcont, recc.cod_cli
),
bens AS (
    SELECT
        ben.nmcont,
        ben.cod_cli,
        CONCAT(ben.marca, ' - ', ben.modelo) AS MarcaModelo,
        ben.placa,
        ben.cor,
        CONCAT(ben.anofab, '/', ben.anomodelo) AS AnoFabModelo,
        COUNT(DISTINCT ben.chassi) AS QtdGarantiasUnicas
    FROM bens_tb ben
    WHERE ben.cod_cli IN ({cod_cli})
    GROUP BY ben.nmcont, ben.cod_cli
),
telefones AS (
    SELECT
        cad.cod_cad AS cod_cad,
        cad.nomecli AS nome,
        cad.cpfcnpj AS cpf,
        cad.nmcont AS nmcont,
        cad.cod_cli AS cod_cli,
        MAX(recc.int_2) AS BindingID,
        DATE_FORMAT(nascto, '%d-%m-%Y') AS DataNascimento,
        cad.infoad AS Portfolio,
        CONCAT(dddfone,telefone) AS telefones,
        ROW_NUMBER() OVER (
            PARTITION BY tel.cod_cad
            ORDER BY FIELD(tel.status, 2, 4, 5, 6, 1, 0), tel.status
        ) AS num
    FROM cadastros_tb cad
    JOIN fones_tb tel ON tel.cod_cad = cad.cod_cad
    LEFT JOIN rec_comp_tb recc
        ON recc.nmcont = cad.nmcont
       AND cad.cod_cli = recc.cod_cli
    WHERE cad.cod_cli IN ({cod_cli})
      AND CONCAT(dddfone, telefone) NOT REGEXP '([0-9])\\1{{5}}'
      AND cad.cod_cad NOT IN(
            SELECT h.cod_cli
            FROM hist_tb h
            LEFT JOIN stcob_tb s ON s.st = h.ocorr
            WHERE h.cod_cli = cad.cod_cad
              AND h.data_at >= CURDATE() - INTERVAL 30 DAY
              AND (s.bsc NOT LIKE '%%sistema%%' OR s.bsc NOT LIKE '' OR s.bsc IS NOT NULL)
              AND h.cod_usu <> '999'
            GROUP BY h.cod_cli
      )
      AND (tel.status IN (2, 4, 5, 6, 1)
           OR (tel.obs NOT LIKE '%%Descon%%' AND tel.obs NOT LIKE '%%incorret%%'))
      AND cad.stcli <> 'INA'
      AND LENGTH(CONCAT(dddfone, telefone)) >= 8
      AND CONCAT(dddfone, telefone) NOT LIKE '%%X%%'
    GROUP BY cad.cod_cad, cad.nomecli, cad.cpfcnpj, cad.nmcont, cad.cod_cli,
             nascto, cad.infoad, dddfone, telefone, tel.status
),
telefones_filtrados AS (
    SELECT * FROM telefones WHERE num <= {tel_limit}
),
telefones_final AS (
    SELECT
        cod_cad, nome, cpf, nmcont, cod_cli, bindingid, datanascimento, portfolio,
        MAX(CASE WHEN num = 1 THEN telefones END) AS Telefone1,
        MAX(CASE WHEN num = 2 THEN telefones END) AS Telefone2,
        MAX(CASE WHEN num = 3 THEN telefones END) AS Telefone3,
        MAX(CASE WHEN num = 4 THEN telefones END) AS Telefone4,
        MAX(CASE WHEN num = 5 THEN telefones END) AS Telefone5,
        MAX(CASE WHEN num = 6 THEN telefones END) AS Telefone6,
        MAX(CASE WHEN num = 7 THEN telefones END) AS Telefone7
    FROM telefones_filtrados
    GROUP BY cod_cad, nome, cpf, nmcont, cod_cli, bindingid, datanascimento, portfolio
)
SELECT
    t.cod_cad,
    t.nome,
    t.cpf,
    t.bindingid,
    t.datanascimento,
    t.portfolio,
    t.Telefone1, t.Telefone2, t.Telefone3, t.Telefone4, t.Telefone5, t.Telefone6, t.Telefone7,
    b.MarcaModelo,
    b.placa,
    b.cor,
    b.AnoFabModelo,
    b.QtdGarantiasUnicas,
    v.Contratos,
    v.TipoProduto,
    a.UltimoValorAcordado,
    a.TipoAcordo,
    a.DataCriacaoGecobi,
    a.StatusUltimoAcordo
FROM telefones_final t
LEFT JOIN bens b
    ON t.nmcont = b.nmcont AND t.cod_cli = b.cod_cli
JOIN valores v
    ON t.nmcont = v.nmcont AND t.cod_cli = v.cod_cli
LEFT JOIN acordos_ultimos a
    ON t.nmcont = a.nmcont
LEFT JOIN cpc_ultimo cpc
    ON t.nmcont = cpc.nmcont
WHERE t.cod_cli IN ({cod_cli});
"""

SQL_QUEBRAS_REJEITADAS = r"""
WITH
cpc_ultimo AS (
    SELECT
        st.bsc,
        cad.nmcont,
        MAX(his.data_at) AS dt_ultimo_cpc
    FROM cadastros_tb cad
    JOIN hist_tb his ON his.cod_cli = cad.cod_cad
    JOIN stcob_tb st ON st.st = his.ocorr
    WHERE cad.cod_cli IN ({cod_cli})
      AND cad.stcli <> 'INA'
      {infoad_filter}
    GROUP BY cad.nmcont
),
acordos_ranked AS (
    SELECT
        a.nmcont,
        a.cod_aco,
        a.data_aco,
        a.data_cad,
        a.vlr_aco,
        a.qtd_p_aco,
        a.staco,
        ROW_NUMBER() OVER (PARTITION BY a.nmcont ORDER BY a.cod_aco DESC) AS rn_aco
    FROM acordos_tb a
    WHERE a.cod_cli IN ({cod_cli})
      AND a.data_cad >= '2025-07-01'
),
acordos_pagos AS (
    SELECT nmcont
    FROM acordos_ranked
    WHERE rn_aco = 1
      AND staco IN ('P','G','A')
),
acordos_ultimos AS (
    SELECT
        aco.nmcont,
        aco.vlr_aco AS UltimoValorAcordado,
        CASE
            WHEN aco.qtd_p_aco = 1 THEN 'AVISTA'
            WHEN aco.qtd_p_aco > 1 THEN 'PARCELADO'
            ELSE NULL
        END AS TipoAcordo,
        DATE_FORMAT(aco.data_cad, '%d-%m-%Y') AS DataCriacaoGecobi,
        CASE aco.staco
            WHEN 'A' THEN 'Em Acordo'
            WHEN 'E' THEN 'Exceção Rejeitada'
            WHEN 'G' THEN 'Pago'
            WHEN 'Q' THEN 'Quebrado'
            WHEN 'P' THEN 'Em Promessa'
            ELSE 'Sem Dados'
        END AS StatusUltimoAcordo,
        aco.staco
    FROM acordos_ranked aco
    WHERE aco.rn_aco = 1
      AND aco.staco IN ('Q','E','P','G','A')
),
acordos_ex AS (
    SELECT aco.nmcont
    FROM acordos_ranked aco
    WHERE aco.rn_aco = 1
      AND aco.staco IN ('Q','E')
),
valores AS (
    SELECT
        recc.nmcont,
        recc.cod_cli,
        GROUP_CONCAT(DISTINCT rec.fat_parc ORDER BY rec.fat_parc SEPARATOR ' || ') AS Contratos,
        GROUP_CONCAT(DISTINCT recc.char_5 ORDER BY recc.char_5 SEPARATOR ' || ') AS TipoProduto
    FROM rec_comp_tb recc
    LEFT JOIN receber_tb rec
        ON rec.nmcont = recc.nmcont
       AND rec.cod_cli = recc.cod_cli
    WHERE recc.cod_cli IN ({cod_cli})
      AND rec.fat_parc NOT LIKE '%ENTRADA%'
      AND recc.nmcont NOT IN (SELECT nmcont FROM acordos_pagos)
    GROUP BY recc.nmcont, recc.cod_cli
    {vlrparc_having}
),
bens AS (
    SELECT
        ben.nmcont,
        ben.cod_cli,
        CONCAT(ben.marca, ' - ', ben.modelo) AS MarcaModelo,
        ben.placa,
        ben.cor,
        CONCAT(ben.anofab, '/', ben.anomodelo) AS AnoFabModelo,
        COUNT(DISTINCT ben.chassi) AS QtdGarantiasUnicas
    FROM bens_tb ben
    WHERE ben.cod_cli IN ({cod_cli})
    GROUP BY ben.nmcont, ben.cod_cli
),
telefones AS (
    SELECT
        cad.cod_cad AS cod_cad,
        cad.nomecli AS nome,
        cad.cpfcnpj AS cpf,
        cad.nmcont AS nmcont,
        cad.cod_cli AS cod_cli,
        MAX(recc.int_2) AS BindingID,
        DATE_FORMAT(nascto, '%d-%m-%Y') AS DataNascimento,
        cad.infoad AS Portfolio,
        CONCAT(dddfone,telefone) AS telefones,
        ROW_NUMBER() OVER (
            PARTITION BY tel.cod_cad
            ORDER BY FIELD(tel.status, 2, 4, 5, 6, 1, 0), tel.status
        ) AS num
    FROM cadastros_tb cad
    JOIN fones_tb tel
        ON tel.cod_cad = cad.cod_cad
    LEFT JOIN rec_comp_tb recc
        ON recc.nmcont = cad.nmcont AND cad.cod_cli = recc.cod_cli
    WHERE cad.cod_cli IN ({cod_cli})
      AND (tel.status IN (2, 4, 5, 6, 1)
           OR (tel.obs NOT LIKE '%Descon%' AND tel.obs NOT LIKE '%incorret%'))
      AND cad.stcli <> 'INA'
      AND CONCAT(dddfone, telefone) NOT REGEXP '([0-9])\\1{{5}}'
      AND LENGTH(CONCAT(dddfone, telefone)) >= 8
      AND CONCAT(dddfone, telefone) NOT LIKE '%X%'
      {infoad_filter}
    GROUP BY cad.cod_cad, cad.nomecli, cad.cpfcnpj, cad.nmcont, cad.cod_cli,
             nascto, cad.infoad, dddfone, telefone, tel.status
),
telefones_filtrados AS (
    SELECT * FROM telefones WHERE num <= {tel_limit}
),
telefones_final AS (
    SELECT
        cod_cad, nome, cpf, nmcont, cod_cli, bindingid, datanascimento, portfolio,
        MAX(CASE WHEN num = 1 THEN telefones END) AS Telefone1,
        MAX(CASE WHEN num = 2 THEN telefones END) AS Telefone2,
        MAX(CASE WHEN num = 3 THEN telefones END) AS Telefone3,
        MAX(CASE WHEN num = 4 THEN telefones END) AS Telefone4,
        MAX(CASE WHEN num = 5 THEN telefones END) AS Telefone5,
        MAX(CASE WHEN num = 6 THEN telefones END) AS Telefone6,
        MAX(CASE WHEN num = 7 THEN telefones END) AS Telefone7
    FROM telefones_filtrados
    GROUP BY cod_cad, nome, cpf, nmcont, cod_cli, bindingid, datanascimento, portfolio
)
SELECT
    t.cod_cad,
    t.nome,
    t.cpf,
    t.bindingid,
    t.datanascimento,
    t.portfolio,
    t.Telefone1, t.Telefone2, t.Telefone3, t.Telefone4, t.Telefone5, t.Telefone6, t.Telefone7,
    b.MarcaModelo,
    b.placa,
    b.cor,
    b.AnoFabModelo,
    b.QtdGarantiasUnicas,
    v.Contratos,
    v.TipoProduto,
    a.UltimoValorAcordado,
    a.TipoAcordo,
    a.DataCriacaoGecobi,
    a.StatusUltimoAcordo
FROM telefones_final t
LEFT JOIN bens b
    ON t.nmcont = b.nmcont AND t.cod_cli = b.cod_cli
JOIN valores v
    ON t.nmcont = v.nmcont AND t.cod_cli = v.cod_cli
LEFT JOIN acordos_ultimos a
    ON t.nmcont = a.nmcont
JOIN acordos_ex ex
    ON t.nmcont = ex.nmcont
LEFT JOIN cpc_ultimo cpc
    ON t.nmcont = cpc.nmcont
WHERE t.cod_cli IN ({cod_cli});
"""

SQL_GARANTIAS_BENS = r"""
SELECT
    cad.cod_cad,
    cad.nomecli AS nome,
    cad.cpfcnpj AS cpf,
    cad.nmcont,
    cad.cod_cli,
    COUNT(DISTINCT ben.chassi) AS QtdGarantiasUnicas,
    GROUP_CONCAT(DISTINCT CONCAT(ben.marca,' - ',ben.modelo) ORDER BY ben.marca, ben.modelo SEPARATOR ' || ') AS MarcaModelo,
    GROUP_CONCAT(DISTINCT ben.placa ORDER BY ben.placa SEPARATOR ' || ') AS Placas
FROM cadastros_tb cad
JOIN bens_tb ben
    ON ben.nmcont = cad.nmcont
   AND ben.cod_cli = cad.cod_cli
WHERE cad.cod_cli IN ({cod_cli})
  AND cad.stcli <> 'INA'
GROUP BY cad.cod_cad, cad.nomecli, cad.cpfcnpj, cad.nmcont, cad.cod_cli
ORDER BY QtdGarantiasUnicas DESC;
"""

SQL_MAIORES_DIVIDAS = r"""
SELECT
    cad.cod_cad,
    cad.nomecli,
    cad.cpfcnpj,
    cad.nmcont,
    cad.cod_cli,
    COUNT(DISTINCT rec.fat_parc) AS QtdContratos,
    SUM(rec.vlrparc) AS ValorTotalDivida
FROM cadastros_tb cad
JOIN receber_tb rec
    ON rec.nmcont = cad.nmcont
   AND rec.cod_cli = cad.cod_cli
WHERE cad.cod_cli IN ({cod_cli})
  AND cad.stcli <> 'INA'
GROUP BY
    cad.cod_cad,
    cad.nomecli,
    cad.cpfcnpj,
    cad.nmcont,
    cad.cod_cli
{having_filter}
ORDER BY ValorTotalDivida DESC
LIMIT 50;
"""


# =========================
# Mapa de consultas (UI)
# =========================
QUERIES = {
    "Email (nome, CPF/CNPJ, email)": (SQL_EMAIL, "base_email.xlsx", "Email"),
    "Nome + CPF/CNPJ": (SQL_NOME_CPF, "base_nome_cpf.xlsx", "NomeCPF"),
    "Telefones + Melhor Contato (Top 7)": (SQL_TELEFONES_MELHOR_CONTATO, "base_telefones_melhor_contato.xlsx", "TelefonesTop7"),
    "Acordos (Promessa/Em Acordo) P/A": (SQL_ACORDOS_PA, "base_acordos_PA.xlsx", "AcordosPA"),
    "CPC por Periodo (datas)": (SQL_CPC_PERIODO, "base_cpc_periodo.xlsx", "CPCPeriodo"),
    "Sem Historico (ultimos 30 dias)": (SQL_SEM_HIST_30D, "base_sem_hist_30d.xlsx", "SemHist30d"),
    "Garantias (bens_tb)": (SQL_GARANTIAS_BENS, "base_garantias_bens.xlsx", "Garantias"),
    "Quebras Rejeitadas": (SQL_QUEBRAS_REJEITADAS, "base_quebras_rejeitadas.xlsx", "QuebrasRejeitadas"),
    "Nunca Contatados": (SQL_NUNCA, "base_nunca.xlsx", "Nunca"),
    "Base Recentes": (SQL_RECENTES, "base_recentes.xlsx", "Recentes"),
    "Maiores Dividas (valor minimo)": (SQL_MAIORES_DIVIDAS, "base_maiores_dividas.xlsx", "MaioresDividas"),
}


# =========================
# UI
# =========================
class App(tk.Tk):
    def __init__(self):
        super().__init__()

        self.title("Gerador de Bases - Itapeva")
        self.geometry("1040x560")
        self.minsize(980, 540)

        self._build_style()

        self.columnconfigure(0, weight=1)
        self.rowconfigure(1, weight=1)
        self.rowconfigure(2, weight=0)

        self._build_header()
        self._build_body()

        self._refresh_params_visibility()
        self._set_busy(False)

    def _on_query_listbox_change(self, event):
        sel = event.widget.curselection()
        if not sel:
            return
        value = event.widget.get(sel[0])
        self.query_var.set(value)
        self._refresh_params_visibility()

    # -------------------------
    # Style
    # -------------------------
    def _build_style(self):
        style = ttk.Style(self)
        try:
            style.theme_use("clam")
        except Exception:
            pass

        style.configure("App.TFrame", background="#f6f7fb")
        style.configure("Header.TFrame", background="#ffffff")
        style.configure("HeaderTitle.TLabel", font=("Segoe UI", 16, "bold"), background="#ffffff")
        style.configure("HeaderSub.TLabel", font=("Segoe UI", 9), background="#ffffff", foreground="#5a6472")

        style.configure("Card.TLabelframe", padding=12, background="#f6f7fb")
        style.configure("Card.TLabelframe.Label", font=("Segoe UI", 10, "bold"))
        style.configure("CardInner.TFrame", background="#f6f7fb")

        style.configure("Primary.TButton", font=("Segoe UI", 10, "bold"), padding=(14, 8))
        style.configure("Secondary.TButton", font=("Segoe UI", 10), padding=(12, 8))

        style.configure("TCheckbutton", background="#f6f7fb")
        style.configure("TLabel", background="#f6f7fb")
        style.configure("TCombobox", padding=6)

        style.configure("Hint.TLabel", font=("Segoe UI", 9), foreground="#6b7585", background="#f6f7fb")

    # -------------------------
    # Header
    # -------------------------
    def _build_header(self):
        header = ttk.Frame(self, style="Header.TFrame", padding=(16, 14))
        header.grid(row=0, column=0, sticky="ew")
        header.columnconfigure(0, weight=1)

        ttk.Label(header, text="Gerador de Bases", style="HeaderTitle.TLabel").grid(row=0, column=0, sticky="w")
        ttk.Label(
            header,
            text="Dúvidas ou não achou a base? Contato: juridico577@oliveiraeantunes.com.br                                                                                                                             Versão 1.2.2 Beta Teste",
            style="HeaderSub.TLabel",
        ).grid(row=1, column=0, sticky="w", pady=(4, 0))

        ttk.Separator(header, orient="horizontal").grid(row=2, column=0, sticky="ew", pady=(12, 0))

    # -------------------------
    # Body
    # -------------------------
    def _build_body(self):
        body = ttk.Frame(self, style="App.TFrame", padding=(16, 14))
        body.grid(row=1, column=0, sticky="nsew")
        body.columnconfigure(0, weight=0)
        body.columnconfigure(1, weight=1)

        self._build_sidebar(body)
        self._build_content(body)

        # Rodapé
        footer = ttk.Frame(self, style="App.TFrame", padding=(16, 6))
        footer.grid(row=2, column=0, sticky="ew")
        footer.columnconfigure(0, weight=1)

        ttk.Label(
            footer,
            text="© 2026 – Desenvolvido por: Lucas Shimazaki e Natã Rafael DJUR5",
            font=("Segoe UI", 9),
            foreground="#6b7585",
            background="#f6f7fb"
        ).grid(row=0, column=0, sticky="w")

    def _build_sidebar(self, parent):
        sidebar = ttk.Frame(parent, style="App.TFrame")
        sidebar.grid(row=0, column=0, sticky="nsw", padx=(0, 12))
        sidebar.columnconfigure(0, weight=1)

        lf_cart = ttk.LabelFrame(sidebar, text="Carteiras", style="Card.TLabelframe")
        lf_cart.grid(row=0, column=0, sticky="new")
        lf_cart.columnconfigure(0, weight=1)

        self.carteira_vars = []
        self.carteira_checks = []
        for i, (label, code) in enumerate(CARTEIRAS):
            var = tk.BooleanVar(value=False)
            self.carteira_vars.append((var, code))
            cb = ttk.Checkbutton(lf_cart, text=label, variable=var)
            cb.grid(row=i, column=0, sticky="w", pady=4)
            self.carteira_checks.append(cb)

        ttk.Label(sidebar, text=f"Telefone fixo: top {TEL_LIMIT_FIXO}", style="Hint.TLabel").grid(
            row=1, column=0, sticky="w", pady=(10, 0)
        )

    def _build_content(self, parent):
        content = ttk.Frame(parent, style="App.TFrame")
        content.grid(row=0, column=1, sticky="nsew")
        content.columnconfigure(0, weight=1)

        # Consulta
        card_top = ttk.LabelFrame(content, text="Consulta", style="Card.TLabelframe")
        card_top.grid(row=0, column=0, sticky="ew")
        card_top.columnconfigure(0, weight=1)

        inner = ttk.Frame(card_top, style="CardInner.TFrame")
        inner.grid(row=0, column=0, sticky="ew")
        inner.columnconfigure(0, weight=1)

        # Lista de consultas
        self.query_var = tk.StringVar()

        self.query_listbox = tk.Listbox(
            inner,
            height=len(QUERIES),
            exportselection=False
        )
        for q in QUERIES.keys():
            self.query_listbox.insert(tk.END, q)

        self.query_listbox.grid(row=0, column=0, sticky="ew", pady=(0, 8))

        # seleciona a primeira automaticamente
        self.query_listbox.selection_set(0)
        self.query_listbox.see(0)
        self.query_var.set(self.query_listbox.get(0))

        self.query_listbox.bind("<<ListboxSelect>>", self._on_query_listbox_change)

        # Parâmetros
        self.lf_params = ttk.LabelFrame(content, text="Parâmetros", style="Card.TLabelframe")
        self.lf_params.grid(row=1, column=0, sticky="ew", pady=(12, 0))
        self.lf_params.columnconfigure(0, weight=1)

        params_row = ttk.Frame(self.lf_params, style="CardInner.TFrame")
        params_row.grid(row=0, column=0, sticky="ew")

        # CPC por período
        ttk.Label(params_row, text="Data Início (YYYY-MM-DD):").grid(row=0, column=0, sticky="w")
        self.dt_ini_var = tk.StringVar(value="")
        self.dt_ini_entry = ttk.Entry(params_row, textvariable=self.dt_ini_var, width=16)
        self.dt_ini_entry.grid(row=0, column=1, sticky="w", padx=(8, 18))

        ttk.Label(params_row, text="Data Fim (YYYY-MM-DD):").grid(row=0, column=2, sticky="w")
        self.dt_fim_var = tk.StringVar(value="")
        self.dt_fim_entry = ttk.Entry(params_row, textvariable=self.dt_fim_var, width=16)
        self.dt_fim_entry.grid(row=0, column=3, sticky="w", padx=(8, 18))

        self.lbl_dt_hint = ttk.Label(params_row, text="(vazio = sem filtro)", style="Hint.TLabel")
        self.lbl_dt_hint.grid(row=0, column=4, sticky="w")

        # Maiores Dívidas: valor mínimo
        ttk.Label(params_row, text="Valor mínimo da dívida:").grid(row=1, column=0, sticky="w", pady=(10, 0))
        self.min_div_var = tk.StringVar(value="")
        self.min_div_entry = ttk.Entry(params_row, textvariable=self.min_div_var, width=16)
        self.min_div_entry.grid(row=1, column=1, sticky="w", padx=(8, 18), pady=(10, 0))

        self.lbl_min_hint = ttk.Label(params_row, text="Ex.: 10000 ou 10.000,00", style="Hint.TLabel")
        self.lbl_min_hint.grid(row=1, column=2, columnspan=3, sticky="w", pady=(10, 0))

        # Ações + progresso
        actions = ttk.Frame(content, style="App.TFrame")
        actions.grid(row=2, column=0, sticky="ew", pady=(14, 0))
        actions.columnconfigure(3, weight=1)

        self.btn_generate = ttk.Button(actions, text="Gerar Excel", style="Primary.TButton", command=self.gerar_excel)
        self.btn_generate.grid(row=0, column=0, sticky="w")

        self.btn_clear = ttk.Button(actions, text="Limpar seleção", style="Secondary.TButton", command=self.limpar)
        self.btn_clear.grid(row=0, column=1, sticky="w", padx=(10, 0))

        self.progress = ttk.Progressbar(actions, mode="indeterminate")
        self.progress.grid(row=0, column=3, sticky="ew", padx=(14, 0))

    # -------------------------
    # UI State
    # -------------------------
    def _set_busy(self, busy: bool):
        state = "disabled" if busy else "normal"

        self.btn_generate.configure(state=state)
        self.btn_clear.configure(state=state)

        if busy:
            self.query_listbox.configure(state="disabled")
        else:
            self.query_listbox.configure(state="normal")

        for cb in self.carteira_checks:
            cb.configure(state=state)

        if busy:
            self.dt_ini_entry.configure(state="disabled")
            self.dt_fim_entry.configure(state="disabled")
            self.min_div_entry.configure(state="disabled")
            self.progress.start(12)
        else:
            self.progress.stop()
            self._refresh_params_visibility()

    def _refresh_params_visibility(self):
        q = self.query_var.get()

        is_cpc = (q == "CPC por Periodo (datas)")
        is_maiores = (q == "Maiores Dividas (valor minimo)")

        self.dt_ini_entry.configure(state=("normal" if is_cpc else "disabled"))
        self.dt_fim_entry.configure(state=("normal" if is_cpc else "disabled"))
        self.min_div_entry.configure(state=("normal" if is_maiores else "disabled"))

    # -------------------------
    # Actions
    # -------------------------
    def limpar(self):
        # (Opção 7) reset visual também no Listbox
        for var, _ in self.carteira_vars:
            var.set(False)

        self.dt_ini_var.set("")
        self.dt_fim_var.set("")
        self.min_div_var.set("")

        self.query_listbox.selection_clear(0, tk.END)
        self.query_listbox.selection_set(0)
        self.query_listbox.see(0)
        self.query_var.set(self.query_listbox.get(0))

        self._refresh_params_visibility()

    def _build_extra_for_selected_query(self, query_name: str) -> dict:
        # CPC por período
        if query_name == "CPC por Periodo (datas)":
            dt_ini = self.dt_ini_var.get().strip()
            dt_fim = self.dt_fim_var.get().strip()

            extra = {"_tail_params": [], "hist_cad_ref_col": HIST_CAD_REF_COL}

            if dt_ini:
                if not _is_valid_ymd(dt_ini):
                    raise ValueError("Data Início inválida. Use YYYY-MM-DD.")
                extra["dt_ini_filter"] = "AND his.data_at >= %s"
                extra["_tail_params"].append(dt_ini + " 00:00:00")
            else:
                extra["dt_ini_filter"] = ""

            if dt_fim:
                if not _is_valid_ymd(dt_fim):
                    raise ValueError("Data Fim inválida. Use YYYY-MM-DD.")
                extra["dt_fim_filter"] = "AND his.data_at <= %s"
                extra["_tail_params"].append(dt_fim + " 23:59:59")
            else:
                extra["dt_fim_filter"] = ""

            return extra

        # Maiores Dívidas
        if query_name == "Maiores Dividas (valor minimo)":
            min_txt = self.min_div_var.get().strip()
            min_value = _parse_money_br_or_plain(min_txt)

            extra = {"_tail_params": []}
            extra["having_filter"] = "HAVING SUM(rec.vlrparc) >= %s"
            extra["_tail_params"].append(min_value)
            return extra

        return {}

    def gerar_excel(self):
        carteiras = [c for v, c in self.carteira_vars if v.get()]
        if not carteiras:
            messagebox.showwarning("Atenção", "Selecione ao menos uma carteira (517/518/519).")
            return

        query_name = self.query_var.get()
        sql_template, default_filename, sheet_name = QUERIES[query_name]

        path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            initialfile=default_filename,
            filetypes=[("Excel", "*.xlsx")],
            title="Salvar Excel"
        )
        if not path:
            return

        try:
            extra = self._build_extra_for_selected_query(query_name)
        except Exception as e:
            messagebox.showerror("Erro", str(e))
            return

        self._set_busy(True)

        t = threading.Thread(
            target=self._job_gerar_excel,
            args=(sql_template, carteiras, extra, path, sheet_name),
            daemon=True
        )
        t.start()

    def _job_gerar_excel(self, sql_template, carteiras, extra, path, sheet_name):
        try:
            df = run_query(sql_template, carteiras, extra=extra)

            # (Opção 10) Excel bonitinho
            _write_excel_pretty(df, path, sheet_name)

            self.after(0, self._on_job_success, path, len(df))

        except FileNotFoundError as e:
            self.after(0, self._on_job_error, "Credenciais não encontradas", str(e))

        except mysql.connector.Error as e:
            self.after(0, self._on_job_error, "Erro no banco", f"Falha ao conectar/consultar:\n{e}")

        except Exception as e:
            self.after(0, self._on_job_error, "Erro", str(e))

    def _on_job_success(self, path: str, n_rows: int):
        self._set_busy(False)
        messagebox.showinfo("Sucesso", f"Excel gerado com sucesso!\n\nLinhas: {n_rows}\n\n{path}")

    def _on_job_error(self, title: str, msg: str):
        self._set_busy(False)
        messagebox.showerror(title, msg)


if __name__ == "__main__":
    print("ARQUIVO RODANDO:", os.path.abspath(__file__))
    print("TOTAL QUERIES:", len(QUERIES))
    print("LISTA QUERIES:", list(QUERIES.keys()))
    App().mainloop()
